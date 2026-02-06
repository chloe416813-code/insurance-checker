import streamlit as st
import pandas as pd
import io
import msoffcrypto
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import zipfile
import xlsxwriter

# ================= 0. 系統環境檢查 =================
try:
    import openpyxl
    import msoffcrypto
    import xlsxwriter
except ImportError:
    st.error("🛑 缺少必要套件")
    st.stop()

# ================= 1. 核心邏輯區 (來自您提供的程式碼) =================
# 計算年齡的基準日
REF_DATE = datetime(2025, 10, 20)
# 定義黃色標記
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def parse_roc_birthday(roc_val):
    """ 解析民國年生日，回傳 datetime 物件，若格式錯誤回傳 None """
    if roc_val is None: return None
    s = str(roc_val).strip().replace('\t', '').replace(' ', '')
    if s == '' or s.lower() == 'nan': return None

    # 處理常見分隔符與中文
    s_clean = s.replace('年', '.').replace('月', '.').replace('日', '').replace('-', '.').replace('/', '.')

    parts = []
    if '.' in s_clean:
        parts = s_clean.split('.')
    elif s_clean.isdigit():
        # 純數字處理
        if len(s_clean) == 6: parts = [s_clean[:2], s_clean[2:4], s_clean[4:]]
        elif len(s_clean) == 7: parts = [s_clean[:3], s_clean[3:5], s_clean[5:]]

    try:
        if len(parts) != 3: return None
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        if not (1 <= m <= 12 and 1 <= d <= 31): return None
        return datetime(y + 1911, m, d)
    except:
        return None

def calculate_age(born):
    if born is None: return -1
    return REF_DATE.year - born.year - ((REF_DATE.month, REF_DATE.day) < (born.month, born.day))

def open_excel_with_password(file_content, password):
    """ 嘗試開啟 Excel (支援加密與非加密) """
    file_stream = io.BytesIO(file_content)

    # 1. 先嘗試直接開啟 (假設無加密)
    try:
        wb = openpyxl.load_workbook(file_stream)
        return wb
    except:
        # 開啟失敗，可能是加密檔，重置指標
        file_stream.seek(0)

    # 2. 嘗試用密碼解鎖
    if password:
        try:
            decrypted = io.BytesIO()
            office_file = msoffcrypto.OfficeFile(file_stream)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            wb = openpyxl.load_workbook(decrypted)
            return wb
        except Exception:
            # 解密失敗
            return None

    return None

def process_single_file_logic(filename, content, password):
    """ 
    這是您原本的 process_single_file 函式
    為了配合 Streamlit，微調了 print -> return 結構 
    """
    # 嘗試開啟
    wb = open_excel_with_password(content, password)

    if wb is None:
        return None, {"filename": filename, "status": "Fail", "msg": "無法開啟(密碼錯誤或格式不支援)"}

    ws = wb.active

    # 自動尋找欄位
    col_idx_map = {}
    # 這裡稍微增強一點：避免讀到空行，搜尋前幾行
    header_found = False
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value:
                col_idx_map[str(cell.value)] = cell.column
        if '身分證' in col_idx_map or any('身分證' in str(k) for k in col_idx_map.keys()):
            header_found = True
            break
            
    if not header_found:
         # 如果找不到表頭，回退到第一行嘗試
         col_idx_map = {}
         for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value: col_idx_map[str(cell.value)] = cell.column

    # 關鍵字對應
    id_key = next((k for k in col_idx_map.keys() if '身分證' in k), None)
    birth_key = next((k for k in col_idx_map.keys() if '生日' in k and '民國' in k), None)

    stats = {"filename": filename, "under_15": 0, "adult": 0, "errors": 0, "status": "Success", "msg": "OK"}

    if not id_key or not birth_key:
        return None, {"filename": filename, "status": "Fail", "msg": "找不到關鍵欄位(需有'身分證'與'生日(民國)')"}

    xl_birth_col = col_idx_map[birth_key]
    xl_id_col = col_idx_map[id_key]

    # 逐列檢查並標記
    # 注意：這裡從 min_row=2 開始，假設第一列是表頭。如果您的表頭在第3列，這裡可能要調整
    # 為了保險，我們從表頭所在列的下一列開始
    start_row = 2 
    
    for row in ws.iter_rows(min_row=start_row):
        # 1. 檢查生日
        if xl_birth_col:
            # 防止索引超出範圍 (如果該列是空的)
            if xl_birth_col - 1 < len(row):
                cell_birth = row[xl_birth_col - 1]
                birth_dt = parse_roc_birthday(cell_birth.value)

                if birth_dt is None:
                    cell_birth.fill = YELLOW_FILL # 標記黃底
                    stats["errors"] += 1
                else:
                    age = calculate_age(birth_dt)
                    if 0 <= age < 15:
                        stats["under_15"] += 1
                    elif age >= 15:
                        stats["adult"] += 1

        # 2. 檢查身分證
        if xl_id_col:
            if xl_id_col - 1 < len(row):
                cell_id = row[xl_id_col - 1]
                val_id = str(cell_id.value).strip() if cell_id.value else ""

                # 檢查漏填或長度錯誤
                if not val_id or val_id == 'None' or len(val_id) != 10:
                    cell_id.fill = YELLOW_FILL # 標記黃底
                    stats["errors"] += 1

    # 存檔到記憶體
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, stats

# ================= 2. 分頁功能實作 =================

def run_checker_tab(uploaded_files, password):
    processed_files = []
    summary_report = []
    progress_bar = st.progress(0)
    
    for i, file in enumerate(uploaded_files):
        content = file.read()
        processed_data, stats = process_single_file_logic(file.name, content, password)
        
        summary_report.append(stats)
        if processed_data:
            # 這裡回傳的是 openpyxl 處理完的 output (含黃底，但無密碼)
            processed_files.append((f"已檢查_{file.name}", processed_data.getvalue()))
            
        progress_bar.progress((i + 1) / len(uploaded_files))
        
    return processed_files, summary_report

def run_encryptor_tab(uploaded_files, new_password):
    processed_files = []
    progress_bar = st.progress(0)
    
    for i, file in enumerate(uploaded_files):
        try:
            content = file.read()
            # 1. 讀取檔案
            # 這裡使用 pd.read_excel，它會自動處理大部分格式
            # 如果是從分頁1下載下來的檔案，它是沒有密碼的，可以直接讀
            try:
                df = pd.read_excel(io.BytesIO(content))
            except:
                # 萬一使用者上傳了有密碼的檔案
                st.error(f"❌ {file.name}: 讀取失敗。請確認上傳的是【無密碼】的檔案 (例如從分頁1下載的檔案)。")
                continue
            
            # 2. 加密寫入
            # 使用 xlsxwriter 引擎進行加密
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                # 將資料寫入
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                workbook = writer.book
                
                # 設定密碼
                workbook.set_encryption(new_password)
            
            processed_files.append((f"加密_{file.name}", output.getvalue()))
            
        except Exception as e:
            st.error(f"❌ {file.name} 加密失敗: {e}")
            
        progress_bar.progress((i + 1) / len(uploaded_files))
        
    return processed_files

# ================= 3. 主程式介面 =================

st.set_page_config(page_title="投保名單工具箱", page_icon="🧰")
st.title("🧰 科普列車 - 投保名單工具箱")

tab1, tab2 = st.tabs(["🔍 1. 檢查名單", "🔒 2. 批次加密"])

# --- 分頁 1: 檢查 (完全依照您的程式碼) ---
with tab1:
    st.header("名單檢查工具")
    st.info("功能：讀取 Excel (支援加密) -> 標記黃底 -> 輸出 **無密碼** 檔案。")
    st.caption("請使用此頁面檢查檔案，下載確認沒問題後，再到「分頁 2」進行加密。")
    
    check_pass = st.text_input("輸入解鎖密碼 (若檔案無加密可留空)", type="password", key="p1")
    check_files = st.file_uploader("上傳 Excel", type=['xlsx'], accept_multiple_files=True, key="u1")
    
    if check_files and st.button("🚀 開始檢查", key="b1"):
        results, report = run_checker_tab(check_files, check_pass)
        
        if report:
            # 簡單顯示結果
            df_rep = pd.DataFrame(report)
            st.dataframe(df_rep)
            
        if results:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for fname, data in results:
                    zf.writestr(fname, data)
                
                # 報告
                report_str = f"檢查報告 {datetime.now().strftime('%H:%M')}\n"
                for item in report:
                    if item['status'] == 'Success':
                        report_str += f"{item['filename']}: 未滿15歲:{item['under_15']}, 成人:{item['adult']}, 錯誤:{item['errors']}\n"
                    else:
                        report_str += f"{item['filename']}: {item['msg']}\n"
                zf.writestr("檢查報告.txt", report_str)
                
            st.download_button("📦 下載檢查結果 (ZIP)", zip_buffer.getvalue(), "檢查結果.zip", "application/zip")

# --- 分頁 2: 加密 ---
with tab2:
    st.header("Excel 批次加密")
    st.info("功能：將 **無密碼** 的 Excel 檔案加上密碼保護。")
    
    enc_pass = st.text_input("設定新密碼 (必填)", type="password", key="p2")
    enc_files = st.file_uploader("上傳要加密的 Excel (需無密碼)", type=['xlsx'], accept_multiple_files=True, key="u2")
    
    if enc_files:
        if not enc_pass:
            st.warning("請輸入要設定的密碼！")
        else:
            if st.button("🔒 開始加密", key="b2"):
                enc_results = run_encryptor_tab(enc_files, enc_pass)
                
                if enc_results:
                    st.success(f"成功加密 {len(enc_results)} 個檔案")
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w") as zf:
                        for fname, data in enc_results:
                            zf.writestr(fname, data)
                    
                    st.download_button("📦 下載加密檔案 (ZIP)", zip_buffer.getvalue(), "已加密檔案.zip", "application/zip")
